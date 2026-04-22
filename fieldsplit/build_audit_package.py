import csv
import random
import shutil
from pathlib import Path
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill

CKPT = Path("/home/chris/Documents/chagas/checkpoints/hardened_seed69")
OUT = Path("/home/chris/Documents/chagas/audit_package")
SEED = 20260416
N_TP_CTRL = 20
N_TN_CTRL = 20
THRESHOLD = 0.5


def main() -> None:
    rng = random.Random(SEED)
    data = np.load(CKPT / "test_eval.npz", allow_pickle=True)
    labels = data["labels"].astype(int)
    probs = data["probs"].astype(float)
    if "paths" in data.files:
        paths = [str(p) for p in data["paths"]]
    else:
        import sys

        sys.path.insert(0, str(Path(__file__).parent))
        from train_fieldsplit import ImageFolderWithPaths, DATA_ROOT

        ds = ImageFolderWithPaths(DATA_ROOT / "test")
        paths = [p for _, p in [(None, img[0]) for img in ds.imgs]]
        assert len(paths) == len(
            labels
        ), f"path count {len(paths)} != labels {len(labels)}"
    preds = (probs >= THRESHOLD).astype(int)
    fp_idx = [i for i in range(len(labels)) if preds[i] == 1 and labels[i] == 0]
    fn_idx = [i for i in range(len(labels)) if preds[i] == 0 and labels[i] == 1]
    tp_idx = [i for i in range(len(labels)) if preds[i] == 1 and labels[i] == 1]
    tn_idx = [i for i in range(len(labels)) if preds[i] == 0 and labels[i] == 0]
    print(
        f"Available: TP={len(tp_idx)} TN={len(tn_idx)} FP={len(fp_idx)} FN={len(fn_idx)}"
    )
    tp_ctrl = rng.sample(tp_idx, min(N_TP_CTRL, len(tp_idx)))
    tn_ctrl = rng.sample(tn_idx, min(N_TN_CTRL, len(tn_idx)))
    items = []
    items += [(i, "FP") for i in fp_idx]
    items += [(i, "FN") for i in fn_idx]
    items += [(i, "CTRL_TP") for i in tp_ctrl]
    items += [(i, "CTRL_TN") for i in tn_ctrl]
    rng.shuffle(items)
    img_dir = OUT / "images"
    img_dir.mkdir(parents=True, exist_ok=True)
    for f in img_dir.glob("img_*.png"):
        f.unlink()
    mapping_rows = []
    for rank, (src_i, category) in enumerate(items, start=1):
        new_name = f"img_{rank:03d}.png"
        src = Path(paths[src_i])
        dst = img_dir / new_name
        if not src.exists():
            raise FileNotFoundError(src)
        shutil.copy2(src, dst)
        mapping_rows.append(
            dict(
                reader_filename=new_name,
                audit_category=category,
                original_image=src.name,
                model_prob=f"{probs[src_i]:.4f}",
                model_pred_at_0_5=int(preds[src_i]),
                morais_label=int(labels[src_i]),
                source_path=str(src),
            )
        )
    map_path = OUT / "_mapping_DO_NOT_SHARE.csv"
    with map_path.open("w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(mapping_rows[0].keys()))
        w.writeheader()
        for row in mapping_rows:
            w.writerow(row)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "scores"
    ws.append(["image_id", "parasite_present (Yes/No)", "notes_optional"])
    bold = Font(bold=True)
    hdr_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    for c in ws[1]:
        c.font = bold
        c.fill = hdr_fill
    for row in mapping_rows:
        ws.append([row["reader_filename"], "", ""])
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 42
    wb.save(OUT / "reader_template.xlsx")
    counts = {
        "FP": sum((1 for r in mapping_rows if r["audit_category"] == "FP")),
        "FN": sum((1 for r in mapping_rows if r["audit_category"] == "FN")),
        "CTRL_TP": sum((1 for r in mapping_rows if r["audit_category"] == "CTRL_TP")),
        "CTRL_TN": sum((1 for r in mapping_rows if r["audit_category"] == "CTRL_TN")),
    }
    protocol = f'# Independent re-audit protocol — Morais blood-smear quadrants\n\n## What to do\nYou will receive a folder (`images/`) containing **{len(mapping_rows)} quadrant\nimages** named `img_001.png` through `img_{len(mapping_rows):03d}.png`, and an\nExcel scoring template (`reader_template.xlsx`) pre-populated with the image IDs.\n\nFor **each image**, please:\n\n1. Open the image and examine it.\n2. In the `parasite_present (Yes/No)` column of the spreadsheet, enter\n   **Yes** if, in your best judgment, the image contains at least one *T. cruzi*\n   trypomastigote (whether or not it is fully in focus or completely visible);\n   otherwise enter **No**.\n3. Optionally add free-text notes in the `notes_optional` column (e.g.\n   "partially obscured by RBC", "morphology ambiguous").\n4. Save the spreadsheet and return it.\n\n## Important rules\n- **Please score independently.** Do not consult the other readers, do not\n  look up the original Morais annotation for any image, and do not use any\n  other reference while scoring.\n- **The filenames contain no hidden information.** The numbering is a random\n  shuffle; you cannot tell from `img_NNN.png` whether the image was a model\n  disagreement or a control, nor what the original annotation says.\n- Estimated time: ~30–60 seconds per image, ~60–90 minutes total.\n\n## What this is for\nWe trained a deep-learning classifier on the publicly available Morais *et\nal.* (2022) smartphone blood-smear dataset, under a slide-disjoint train /\ntest split. We want an independent expert judgment of whether each image\ncontains a parasite, so that we can characterize the reliability of the\noriginal dataset annotations. Your scores will be combined across the three\nreaders to produce a consensus label per image; we will then report\naggregate counts with statistical caveats in the manuscript.\n\nPlease do not ask us about the composition of the sample (which images are\nwhich) until after you have completed and returned your scores — any such\ninformation during scoring can bias the result we are trying to measure.\n\nThank you!\n'
    (OUT / "PROTOCOL.md").write_text(protocol)
    print(f"\nPackage built at {OUT}/")
    print(f"  images/                    {len(mapping_rows)} files")
    print(f"  reader_template.xlsx       blank per-reader sheet")
    print(f"  PROTOCOL.md                instructions")
    print(f"  _mapping_DO_NOT_SHARE.csv  private mapping")
    print(f"\nCategory counts:")
    for k, v in counts.items():
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
